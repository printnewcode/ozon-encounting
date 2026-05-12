import io
import json
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import quote, urlencode

import pandas as pd
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_POST
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import SalesReportPeriodForm, SalesUploadForm, SupplyUploadForm
from .models import Product, SaleRecord
from .services.ozon_client import OzonAPIError
from .services.ozon_sync import OzonSyncService


MONEY_HEADERS = ('Стоимость в закупке', 'Доставка', 'Себестоимость', 'Доход', 'Прибыль')
DATE_HEADER = 'Дата продажи'
CSV_ENCODINGS = ('utf-8-sig', 'utf-8', 'cp1251')
MAX_VISIBLE_WARNINGS = 5
PRODUCT_GROUPS_PER_PAGE = 25
PRODUCT_FILTERS = ('sold', 'in_stock', 'in_sale', 'profit_positive', 'profit_negative')
PRODUCT_SORTS = ('article', 'name', 'status', 'cost', 'income', 'profit', 'date', 'accrual_date', 'accrual_id')


def selected_product_filters(request):
    if 'filters_applied' not in request.GET:
        return list(PRODUCT_FILTERS)
    return [value for value in request.GET.getlist('filters') if value in PRODUCT_FILTERS]


def product_list_query(request, **updates) -> str:
    query = request.GET.copy()
    query.pop('page', None)
    for key, value in updates.items():
        query.pop(key, None)
        if value is None:
            continue
        if isinstance(value, (list, tuple)):
            query.setlist(key, [str(item) for item in value])
        else:
            query[key] = str(value)
    return urlencode(list(query.lists()), doseq=True)


def export_filename(name: str) -> str:
    timestamp = datetime.now().strftime('%Y-%m-%d')
    return f'{name}_{timestamp}.xlsx'


def workbook_response(workbook: Workbook, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


def style_export_sheet(sheet, widths: list[int]) -> None:
    header_fill = PatternFill('solid', fgColor='0EA5E9')
    header_font = Font(color='FFFFFF', bold=True)
    border_color = 'E5E7EB'
    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color),
    )

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 22

    for column in sheet.iter_cols():
        header = column[0].value
        for cell in column[1:]:
            if header in MONEY_HEADERS:
                cell.number_format = '#,##0.00'
            elif header == DATE_HEADER:
                cell.number_format = 'DD.MM.YYYY'


def parse_decimal(value, field_name: str, row_number: int) -> Decimal:
    try:
        return Decimal(str(value).strip().replace(',', '.'))
    except (InvalidOperation, ValueError):
        raise ValueError(f"Строка {row_number}: некорректное значение поля «{field_name}»: {value!r}")


def parse_positive_int(value, field_name: str, row_number: int, default: int | None = None) -> int:
    if value in (None, '') and default is not None:
        return default
    try:
        quantity = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"Строка {row_number}: некорректное значение поля «{field_name}»: {value!r}")
    if quantity <= 0:
        raise ValueError(f"Строка {row_number}: поле «{field_name}» должно быть больше 0")
    return quantity


def is_header_or_empty_article(value) -> bool:
    article = str(value).strip()
    return not article or article.lower() in {'nan', 'article', 'артикул'}


def parse_request_decimal(value, field_name: str) -> Decimal:
    try:
        amount = Decimal(str(value).strip().replace(',', '.')).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        raise ValueError(f'Некорректное значение поля «{field_name}».')
    if amount < 0:
        raise ValueError(f'Поле «{field_name}» не может быть меньше 0.')
    return amount


def values_differ(left, right) -> bool:
    return str(left).strip() != str(right).strip()


def add_limited_warning(warnings: list[str], message: str) -> None:
    if message not in warnings:
        warnings.append(message)


def show_import_warnings(request, warnings: list[str]) -> None:
    if not warnings:
        return
    visible_warnings = warnings[:MAX_VISIBLE_WARNINGS]
    suffix = f" Еще предупреждений: {len(warnings) - MAX_VISIBLE_WARNINGS}." if len(warnings) > MAX_VISIBLE_WARNINGS else ''
    messages.warning(request, " ".join(visible_warnings) + suffix)


def sale_cost_price(sale: SaleRecord) -> Decimal:
    if sale.profit is not None:
        return sale.income - sale.profit
    return sale.product.cost_price


def parse_file(file) -> pd.DataFrame:
    """Parse uploaded Excel/CSV file into a normalized DataFrame."""
    if Path(file.name).suffix.lower() == '.csv':
        content = file.read()

        for encoding in CSV_ENCODINGS:
            try:
                decoded_content = content.decode(encoding)
                separator = ';' if ';' in decoded_content else ','
                df = pd.read_csv(io.StringIO(decoded_content), sep=separator)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Не удалось определить кодировку файла. Используйте UTF-8 или CP1251.")
    else:
        df = pd.read_excel(file)

    if df.shape[1] and not is_header_or_empty_article(df.columns[0]):
        first_row = pd.DataFrame([list(df.columns)])
        df = pd.concat([first_row, df], ignore_index=True)
        df.columns = range(df.shape[1])

    df.columns = df.columns.astype(str).str.strip()
    return df.fillna('')


def upload_supply(request):
    if request.method == 'POST':
        form = SupplyUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                df = parse_file(file)

                products_created = 0
                warnings = []
                seen_articles = set()
                with transaction.atomic():
                    for index, row in df.iterrows():
                        row_number = index + 1
                        if df.shape[1] < 5:
                            raise ValueError("Файл поставки должен содержать минимум 5 колонок.")

                        article = str(row.iloc[0]).strip()
                        if is_header_or_empty_article(article):
                            continue
                        if article in seen_articles:
                            add_limited_warning(
                                warnings,
                                f"Артикул {article} встречается в файле поставки несколько раз; количество будет суммировано.",
                            )
                        seen_articles.add(article)

                        name = str(row.iloc[1]).strip()
                        if not name:
                            raise ValueError(f"Строка {row_number}: название товара не может быть пустым")

                        purchase_price = parse_decimal(row.iloc[2], 'Стоимость в закупке', row_number)
                        delivery_cost = parse_decimal(row.iloc[3], 'Доставка', row_number)
                        quantity_value = row.iloc[5] if df.shape[1] > 5 else row.iloc[4]
                        quantity = parse_positive_int(quantity_value, 'Количество', row_number)

                        product, created = Product.objects.select_for_update().get_or_create(
                            article=article,
                            defaults={
                                'name': name,
                                'purchase_price': purchase_price,
                                'delivery_cost': delivery_cost,
                                'quantity': quantity,
                                'status': 'in_stock',
                            },
                        )
                        if not created:
                            if values_differ(product.name, name):
                                add_limited_warning(
                                    warnings,
                                    f"Для артикула {article} название в файле отличается от базы: «{name}» / «{product.name}».",
                                )
                            if product.purchase_price != purchase_price or product.delivery_cost != delivery_cost:
                                add_limited_warning(
                                    warnings,
                                    f"Для артикула {article} изменена цена поставки или доставки; себестоимость пересчитана.",
                                )
                            product.purchase_price = purchase_price
                            product.delivery_cost = delivery_cost
                            product.quantity += quantity
                            product.status = 'in_stock'
                            product.save()

                        products_created += 1

                messages.success(request, f'Успешно загружено {products_created} товаров из поставки.')
                show_import_warnings(request, warnings)
                return redirect('product_list')
            except Exception as exc:
                messages.error(request, f'Ошибка при обработке файла: {exc}')
    else:
        form = SupplyUploadForm()

    return render(request, 'web/upload_supply.html', {'form': form})


def upload_sales(request):
    if request.method == 'POST':
        form = SalesUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            sale_type = form.cleaned_data['sale_type']
            try:
                df = parse_file(file)

                sales_created = 0
                errors = []
                warnings = []
                seen_articles = set()
                with transaction.atomic():
                    for index, row in df.iterrows():
                        row_number = index + 1
                        if df.shape[1] < 3:
                            raise ValueError("Файл продаж должен содержать минимум 3 колонки.")

                        article = str(row.iloc[0]).strip()
                        if is_header_or_empty_article(article):
                            continue
                        if article in seen_articles:
                            add_limited_warning(
                                warnings,
                                f"Артикул {article} встречается в файле продаж несколько раз; продажи будут суммированы.",
                            )
                        seen_articles.add(article)

                        uploaded_name = str(row.iloc[1]).strip() if df.shape[1] > 1 else ''

                        income = parse_decimal(row.iloc[2], 'Доход', row_number)
                        quantity_value = row.iloc[3] if df.shape[1] > 3 else None
                        quantity_sold = parse_positive_int(quantity_value, 'Количество', row_number, default=1)

                        try:
                            product = Product.objects.select_for_update().get(article=article)
                        except Product.DoesNotExist:
                            errors.append(f"Товар с артикулом {article} не найден в базе.")
                            continue

                        if uploaded_name and values_differ(product.name, uploaded_name):
                            add_limited_warning(
                                warnings,
                                f"Для артикула {article} название в файле отличается от базы: «{uploaded_name}» / «{product.name}».",
                            )
                        if quantity_sold > product.quantity:
                            add_limited_warning(
                                warnings,
                                f"По артикулу {article} продано {quantity_sold}, а в остатке было {product.quantity}; остаток обнулен.",
                            )

                        product.quantity -= quantity_sold
                        if product.quantity <= 0:
                            product.status = 'sold'
                            product.quantity = 0
                        elif sale_type == 'ozon':
                            product.status = 'in_sale'
                        product.save()

                        for _ in range(quantity_sold):
                            SaleRecord.objects.create(
                                product=product,
                                sale_type=sale_type,
                                income=income,
                            )
                        sales_created += quantity_sold

                if sales_created > 0:
                    messages.success(request, f'Успешно загружено {sales_created} записей о продажах.')
                if errors:
                    message = "Некоторые товары не были найдены: " + ", ".join(errors[:5])
                    messages.warning(request, message + ("..." if len(errors) > 5 else ""))
                show_import_warnings(request, warnings)

                return redirect('product_list')
            except Exception as exc:
                messages.error(request, f'Ошибка при обработке файла: {exc}')
    else:
        form = SalesUploadForm()

    return render(request, 'web/upload_sales.html', {'form': form})


def product_list(request):
    groups_dict = {}
    active_filters = selected_product_filters(request)
    active_article_query = request.GET.get('article', '').strip()
    active_sort = request.GET.get('sort', 'status')
    if active_sort not in PRODUCT_SORTS:
        active_sort = 'status'
    active_direction = request.GET.get('direction', 'asc')
    if active_direction not in ('asc', 'desc'):
        active_direction = 'asc'
    status_order = {
        'in_stock': 0,
        'in_sale': 1,
        'sold': 2,
    }

    def get_group(article, status_key, name, status_label):
        group_key = (article, status_key)
        if group_key not in groups_dict:
            groups_dict[group_key] = {
                'article': article,
                'name': name,
                'status_key': status_key,
                'status_label': status_label,
                'rows': [],
                'count': 0,
                'total_income': Decimal('0'),
                'total_profit': Decimal('0'),
                'has_sales': False,
                'last_sale_date': None,
                'product_ids': [],
                'sort_status': status_order.get(status_key, 99),
            }
        return groups_dict[group_key]

    unsold = Product.objects.filter(status__in=['in_stock', 'in_sale']).order_by('article', 'created_at')
    for product in unsold:
        group = get_group(product.article, product.status, product.name, product.get_status_display())
        if product.id not in group['product_ids']:
            group['product_ids'].append(product.id)

        if product.status == 'in_stock':
            group['rows'].append({
                'product_id': product.id,
                'article': product.article,
                'name': product.name,
                'status_key': product.status,
                'status_label': product.get_status_display(),
                'cost_price': product.cost_price,
                'income': None,
                'profit': None,
                'sale_date': None,
                'accrual_date': None,
                'accrual_id': '',
                'accrual_details_json': '',
            })
            group['count'] += max(product.quantity, 0)
            continue

        remaining_count = max(product.quantity, 1)
        for _ in range(remaining_count):
            group['rows'].append({
                'product_id': product.id,
                'article': product.article,
                'name': product.name,
                'status_key': product.status,
                'status_label': product.get_status_display(),
                'cost_price': product.cost_price,
                'income': None,
                'profit': None,
                'sale_date': None,
                'accrual_date': None,
                'accrual_id': '',
                'accrual_details_json': '',
            })
            group['count'] += 1

    sales = SaleRecord.objects.select_related('product').order_by('article', 'sale_date', 'created_at')
    for sale in sales:
        group = get_group(sale.article, 'sold', sale.name, 'Продан')

        group['rows'].append({
            'product_id': sale.product_id,
            'article': sale.article,
            'name': sale.name,
            'status_key': 'sold',
            'status_label': 'Продан',
            'cost_price': sale_cost_price(sale),
            'income': sale.income,
            'profit': sale.profit,
            'sale_date': sale.sale_date,
            'accrual_date': sale.accrual_date,
            'accrual_id': sale.accrual_id or '',
            'accrual_details_json': json.dumps(sale.accrual_details, ensure_ascii=False) if sale.accrual_details else '',
        })
        group['count'] += 1
        if sale.income:
            group['total_income'] += sale.income
        if sale.profit:
            group['total_profit'] += sale.profit
        group['has_sales'] = True
        if group['last_sale_date'] is None or sale.sale_date > group['last_sale_date']:
            group['last_sale_date'] = sale.sale_date

    groups = []
    def is_group_visible(group):
        if active_article_query and active_article_query.lower() not in str(group['article']).lower():
            return False
        if group['status_key'] not in active_filters:
            return False
        if group['status_key'] == 'sold':
            if group['total_profit'] < 0:
                return 'profit_negative' in active_filters
            return 'profit_positive' in active_filters
        return True

    def sort_value(group):
        if active_sort == 'article':
            return str(group['article']).lower()
        if active_sort == 'name':
            return str(group['name']).lower()
        if active_sort == 'status':
            return group['sort_status']
        if active_sort == 'cost':
            return group['rows'][0]['cost_price']
        if active_sort == 'income':
            return group['total_income'] if group['has_sales'] else Decimal('-1')
        if active_sort == 'profit':
            return group['total_profit']
        if active_sort == 'date':
            return group['last_sale_date'] or date.min
        if active_sort == 'accrual_date':
            dates = [row['accrual_date'] for row in group['rows'] if row.get('accrual_date')]
            return max(dates) if dates else date.min
        if active_sort == 'accrual_id':
            ids = [str(row['accrual_id']).lower() for row in group['rows'] if row.get('accrual_id')]
            return ids[0] if ids else ''
        return group['sort_status']

    for group in groups_dict.values():
        if not is_group_visible(group):
            continue
        group['header_status_key'] = group['status_key']
        group['header_status_label'] = group['status_label']
        group['header_cost'] = group['rows'][0]['cost_price']
        group['product_ids_value'] = ','.join(str(product_id) for product_id in group['product_ids'])
        group['primary_product_id'] = group['product_ids'][0] if group['product_ids'] else ''
        groups.append(group)

    groups.sort(
        key=lambda item: (sort_value(item), item['sort_status'], str(item['article']).lower(), str(item['name']).lower()),
        reverse=active_direction == 'desc',
    )

    paginator = Paginator(groups, PRODUCT_GROUPS_PER_PAGE)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_query = product_list_query(request)

    table_headers = [
        ('article', 0, 'str', 'Артикул'),
        ('name', 1, 'str', 'Название'),
        ('status', 2, 'str', 'Состояние'),
        ('cost', 3, 'num', 'Себестоимость'),
        ('income', 4, 'num', 'Доход'),
        ('profit', 5, 'num', 'Прибыль'),
        ('date', 6, 'date', 'Дата продажи'),
        ('accrual_date', 7, 'date', 'Дата начисления'),
        ('accrual_id', 8, 'str', 'ID начисления'),
    ]
    table_headers = [
        {
            'key': key,
            'col': col,
            'type': value_type,
            'label': label,
            'direction': active_direction if key == active_sort else '',
            'url': '?' + product_list_query(
                request,
                sort=key,
                direction='desc' if key == active_sort and active_direction == 'asc' else 'asc',
            ),
        }
        for key, col, value_type, label in table_headers
    ]

    return render(request, 'web/product_list.html', {
        'groups': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_groups': paginator.count,
        'active_filters': active_filters,
        'active_article_query': active_article_query,
        'active_sort': active_sort,
        'active_direction': active_direction,
        'page_query_prefix': f'{page_query}&' if page_query else '',
        'table_headers': table_headers,
    })


@require_POST
def update_cost_price(request):
    try:
        product_id = int(request.POST.get('product_id') or 0)
        new_cost_price = parse_request_decimal(request.POST.get('cost_price'), 'Себестоимость')
        apply_to_sales = request.POST.get('apply_to_sales') == 'true'
    except (TypeError, ValueError) as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)

    with transaction.atomic():
        try:
            product = Product.objects.select_for_update().get(id=product_id, status='in_stock')
        except Product.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Товар в наличии не найден.'}, status=404)

        sale_queryset = SaleRecord.objects.select_for_update().filter(article=product.article)
        sales_snapshot = list(sale_queryset.values('id', 'profit'))

        request.session['cost_price_undo'] = {
            'product_id': product.id,
            'article': product.article,
            'purchase_price': str(product.purchase_price),
            'delivery_cost': str(product.delivery_cost),
            'cost_price': str(product.cost_price),
            'sales': [
                {'id': sale['id'], 'profit': str(sale['profit']) if sale['profit'] is not None else None}
                for sale in sales_snapshot
            ],
        }

        product.purchase_price = new_cost_price
        product.delivery_cost = Decimal('0.00')
        product.save()

        updated_sales = 0
        if apply_to_sales:
            for sale in sale_queryset:
                sale.profit = sale.income - new_cost_price
                sale.save(update_fields=['profit'])
                updated_sales += 1

    return JsonResponse({
        'ok': True,
        'article': product.article,
        'cost_price': str(product.cost_price),
        'updated_sales': updated_sales,
    })


@require_POST
def undo_cost_price(request):
    undo_data = request.session.get('cost_price_undo')
    if not undo_data:
        return JsonResponse({'ok': False, 'error': 'Нет изменений для отмены.'}, status=404)

    with transaction.atomic():
        try:
            product = Product.objects.select_for_update().get(id=undo_data['product_id'])
        except Product.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Товар не найден.'}, status=404)

        product.purchase_price = Decimal(undo_data['purchase_price'])
        product.delivery_cost = Decimal(undo_data['delivery_cost'])
        product.save()

        restored_sales = 0
        for sale_data in undo_data.get('sales', []):
            updated = SaleRecord.objects.filter(id=sale_data['id']).update(
                profit=Decimal(sale_data['profit']) if sale_data['profit'] is not None else None,
            )
            restored_sales += updated

        del request.session['cost_price_undo']

    return JsonResponse({
        'ok': True,
        'article': product.article,
        'cost_price': str(product.cost_price),
        'restored_sales': restored_sales,
    })


def sales_statistics(request):
    initial = {'date_to': timezone.localdate()}
    form = SalesReportPeriodForm(request.GET or None, initial=initial)
    date_from = None
    date_to = timezone.localdate()
    has_period = bool(request.GET.get('date_from') or request.GET.get('date_to'))

    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to') or (timezone.localdate() if has_period else None)
    elif request.GET:
        messages.error(request, 'Проверьте даты периода.')

    def build_sales_stats(sales_queryset):
        stats = {
            'sales_count': 0,
            'income': Decimal('0'),
            'profit': Decimal('0'),
            'positive_count': 0,
            'negative_count': 0,
            'avg_profit': Decimal('0'),
        }
        for sale in sales_queryset:
            profit = sale.profit or Decimal('0')
            stats['sales_count'] += 1
            stats['income'] += sale.income or Decimal('0')
            stats['profit'] += profit
            if profit < 0:
                stats['negative_count'] += 1
            else:
                stats['positive_count'] += 1
        if stats['sales_count']:
            stats['avg_profit'] = stats['profit'] / stats['sales_count']
        return stats

    visible_sales = SaleRecord.objects.select_related('product').order_by('sale_date', 'created_at')
    if has_period:
        if date_from:
            visible_sales = visible_sales.filter(sale_date__gte=date_from)
        if date_to:
            visible_sales = visible_sales.filter(sale_date__lte=date_to)

    stock_stats = {
        'in_stock_count': 0,
        'in_sale_count': 0,
        'stock_value': Decimal('0'),
    }
    for product in Product.objects.filter(status__in=['in_stock', 'in_sale']):
        quantity = max(product.quantity, 0)
        if product.status == 'in_stock':
            stock_stats['in_stock_count'] += quantity
        elif product.status == 'in_sale':
            stock_stats['in_sale_count'] += quantity
        stock_stats['stock_value'] += product.cost_price * quantity

    return render(request, 'web/statistics.html', {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'sales_stats': build_sales_stats(visible_sales),
        'sales_stats_title': 'За период' if has_period else 'За все время',
        'has_period': has_period,
        'stock_stats': stock_stats,
    })


@require_POST
def sync_ozon(request):
    date_to = timezone.localdate()
    date_from = date_to - timedelta(days=30)

    try:
        result = OzonSyncService().sync_all(date_from, date_to)
        messages.success(
            request,
            (
                'Синхронизация с Ozon завершена: '
                f'создано товаров: {result.products_created}, '
                f'обновлено товаров: {result.products_updated}, '
                f'обновлено остатков: {result.stocks_updated}, '
                f'создано продаж: {result.sales_created}, '
                f'пропущено дублей: {result.sales_skipped}.'
            ),
        )
    except OzonAPIError as exc:
        messages.error(request, f'Ошибка Ozon API: {exc}')

    return redirect('product_list')


def sales_report_period(request):
    initial = {'date_to': timezone.localdate()}
    form = SalesReportPeriodForm(request.GET or None, initial=initial)
    return render(request, 'web/sales_report_period.html', {'form': form})


def export_sales_report(request):
    form = SalesReportPeriodForm(request.GET)
    if not form.is_valid():
        messages.error(request, 'Проверьте даты отчетного периода.')
        return redirect('sales_report_period')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Отчет продаж'
    sheet.append(['Артикул', 'Название', 'Состояние', 'Себестоимость', 'Доход', 'Прибыль', 'Дата продажи'])

    sales = SaleRecord.objects.select_related('product').order_by('sale_date', 'created_at', 'article')
    date_from = form.cleaned_data.get('date_from')
    date_to = form.cleaned_data.get('date_to')
    if date_from:
        sales = sales.filter(sale_date__gte=date_from)
    if date_to:
        sales = sales.filter(sale_date__lte=date_to)

    if not sales.exists():
        messages.warning(request, 'За выбранный период продаж нет. Отчет не сформирован.')
        return redirect('sales_report_period')

    for sale in sales:
        sheet.append([
            sale.article,
            sale.name,
            'Продан',
            sale_cost_price(sale),
            sale.income,
            sale.profit,
            sale.sale_date,
        ])

    style_export_sheet(sheet, [20, 36, 16, 16, 14, 14, 16])
    return workbook_response(workbook, export_filename('Отчет_продаж'))


def export_stock_balance(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Учет остатков'
    sheet.append([
        'Артикул',
        'Название',
        'Стоимость в закупке',
        'Доставка',
        'Себестоимость',
        'На складе',
        'В продаже',
    ])

    products = Product.objects.order_by('article', 'name')
    for product in products:
        in_stock = product.quantity if product.status == 'in_stock' else 0
        in_sale = product.quantity if product.status == 'in_sale' else 0
        sheet.append([
            product.article,
            product.name,
            product.purchase_price,
            product.delivery_cost,
            product.cost_price,
            in_stock,
            in_sale,
        ])

    style_export_sheet(sheet, [20, 36, 20, 14, 16, 14, 14])
    return workbook_response(workbook, export_filename('Учет_остатков'))
