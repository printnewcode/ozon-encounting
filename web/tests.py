from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from django.urls import reverse
from openpyxl import load_workbook

from .models import Product, SaleRecord
from .services.ozon_sync import OzonSyncService
from .views import parse_file


class FakeOzonClient:
    def __init__(self, finance_transactions=None, product_stocks=None, stock_on_warehouses=None):
        self._finance_transactions = finance_transactions
        self._product_stocks = product_stocks
        self._stock_on_warehouses = stock_on_warehouses
        self.finance_transaction_periods = []

    def product_list(self):
        return iter([
            {'offer_id': 'OZON-1', 'product_id': 101, 'visibility': 'VISIBLE'},
        ])

    def product_info_list(self, offer_ids):
        return [
            {'offer_id': 'OZON-1', 'id': 101, 'sku': 202, 'name': 'Товар Ozon', 'statuses': {'status': 'ready'}},
        ]

    def product_stocks(self):
        if self._product_stocks is not None:
            return iter(self._product_stocks)
        return iter([
            {'offer_id': 'OZON-1', 'visibility': 'VISIBLE', 'stocks': [{'present': 3}, {'present': 2}]},
        ])

    def stock_on_warehouses(self):
        if self._stock_on_warehouses is not None:
            return iter(self._stock_on_warehouses)
        return iter([])

    def fbo_postings(self, date_from, date_to):
        return iter([
            {
                'posting_number': 'FBO-1',
                'in_process_at': '2026-04-20T10:00:00Z',
                'products': [
                    {'offer_id': 'OZON-1', 'name': 'Товар Ozon', 'price': '150.50', 'quantity': 2},
                ],
            },
        ])

    def fbs_postings(self, date_from, date_to):
        return iter([
            {
                'posting_number': 'FBS-1',
                'shipment_date': '2026-04-21T10:00:00Z',
                'products': [
                    {'offer_id': 'OZON-1', 'name': 'Товар Ozon', 'price': '200.00', 'quantity': 1},
                ],
            },
        ])

    def finance_transactions(self, date_from, date_to):
        self.finance_transaction_periods.append((date_from, date_to))
        if self._finance_transactions is None:
            return iter([
                {
                    'operation_id': 10001,
                    'operation_date': '2026-04-22T12:00:00Z',
                    'posting': {'posting_number': 'FBO-1'},
                    'amount': '241.00',
                    'items': [{'sku': 202, 'accruals_for_sale': '301.00'}],
                    'services': [{'name': 'MarketplaceServiceItemDirectFlowLogistic', 'price': '-30.00'}],
                },
                {
                    'operation_id': 10002,
                    'operation_date': '2026-04-23T12:00:00Z',
                    'posting': {'posting_number': 'FBS-1'},
                    'amount': '150.00',
                    'items': [{'sku': 202, 'accruals_for_sale': '200.00'}],
                    'services': [{'name': 'commission', 'price': '-50.00'}],
                },
            ])
        return iter(self._finance_transactions)


class ProductModelTests(TestCase):
    def test_save_calculates_cost_price(self):
        product = Product.objects.create(
            article='SKU-1',
            name='Тестовый товар',
            quantity=2,
            purchase_price=Decimal('100.50'),
            delivery_cost=Decimal('20.25'),
        )

        self.assertEqual(product.cost_price, Decimal('120.75'))


class OzonSyncTests(TestCase):
    def setUp(self):
        self.service = OzonSyncService(client=FakeOzonClient())

    def test_sync_products_creates_product_from_ozon(self):
        result = self.service.sync_products()

        product = Product.objects.get(article='OZON-1')
        self.assertEqual(result.products_created, 1)
        self.assertEqual(product.name, 'Товар Ozon')
        self.assertEqual(product.ozon_product_id, 101)
        self.assertEqual(product.ozon_sku, 202)
        self.assertEqual(product.ozon_visibility, 'VISIBLE')
        self.assertEqual(product.ozon_status, 'ready')
        self.assertEqual(product.purchase_price, Decimal('0.00'))

    def test_sync_stocks_updates_quantity_and_status(self):
        Product.objects.create(
            article='OZON-1',
            name='Товар Ozon',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
        )

        result = self.service.sync_stocks()

        product = Product.objects.get(article='OZON-1')
        self.assertEqual(result.stocks_updated, 1)
        self.assertEqual(product.quantity, 5)
        self.assertEqual(product.status, 'in_sale')

    def test_sync_stocks_keeps_blocked_product_out_of_sale(self):
        product = Product.objects.create(
            article='OZON-1',
            name='Товар Ozon',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            ozon_visibility='INVISIBLE',
            ozon_status='validation_failed',
        )

        result = self.service.sync_stocks()

        product.refresh_from_db()
        self.assertEqual(result.stocks_updated, 1)
        self.assertEqual(product.quantity, 5)
        self.assertEqual(product.status, 'in_stock')

    def test_sync_stocks_uses_warehouse_analytics_when_product_stocks_are_empty(self):
        client = FakeOzonClient(
            product_stocks=[
                {'offer_id': 'OZON-1', 'visibility': '', 'stocks': [{'present': 0}]},
            ],
            stock_on_warehouses=[
                {'item_code': 'OZON-1', 'free_to_sell_amount': 15},
            ],
        )
        service = OzonSyncService(client=client)
        Product.objects.create(
            article='OZON-1',
            name='РўРѕРІР°СЂ Ozon',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
        )

        result = service.sync_stocks()

        product = Product.objects.get(article='OZON-1')
        self.assertEqual(result.stocks_updated, 1)
        self.assertEqual(product.quantity, 15)
        self.assertEqual(product.status, 'in_sale')

    def test_sync_postings_creates_sales_without_duplicates(self):
        product = Product.objects.create(
            article='OZON-1',
            name='Товар Ozon',
            quantity=5,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
        )

        first_result = self.service.sync_postings(date(2026, 4, 1), date(2026, 4, 28))
        second_result = self.service.sync_postings(date(2026, 4, 1), date(2026, 4, 28))

        self.assertEqual(first_result.sales_created, 3)
        self.assertEqual(second_result.sales_created, 0)
        self.assertEqual(second_result.sales_skipped, 3)
        self.assertEqual(SaleRecord.objects.count(), 3)
        self.assertEqual(SaleRecord.objects.filter(posting_number='FBO-1').count(), 2)
        self.assertEqual(SaleRecord.objects.filter(posting_number='FBS-1').count(), 1)
        self.assertEqual(SaleRecord.objects.filter(sale_date=date(2026, 4, 20)).count(), 2)
        self.assertEqual(SaleRecord.objects.filter(sale_date=date(2026, 4, 21)).count(), 1)
        self.assertEqual(SaleRecord.objects.filter(accrual_id='10001', accrual_date=date(2026, 4, 22)).count(), 2)
        self.assertEqual(SaleRecord.objects.filter(accrual_id='10002', accrual_date=date(2026, 4, 23)).count(), 1)
        self.assertEqual(SaleRecord.objects.first().product, product)
        self.assertEqual(
            list(SaleRecord.objects.order_by('external_id').values_list('income', flat=True)),
            [Decimal('120.50'), Decimal('120.50'), Decimal('150.00')],
        )
        self.assertEqual(
            list(SaleRecord.objects.order_by('external_id').values_list('accrual_id', flat=True)),
            ['10001', '10001', '10002'],
        )
        first_sale = SaleRecord.objects.order_by('external_id').first()
        self.assertEqual(first_sale.accrual_details['gross_price'], '150.50')
        self.assertEqual(first_sale.accrual_details['net_income'], '120.50')
        self.assertEqual(first_sale.accrual_details['deductions_total'], '30.00')
        self.assertEqual(first_sale.accrual_details['services'][0]['name'], 'Логистика до покупателя')
        self.assertEqual(first_sale.accrual_details['services'][0]['code'], 'MarketplaceServiceItemDirectFlowLogistic')
        self.assertEqual(first_sale.accrual_details['services'][0]['price'], '-30.00')
        self.assertEqual(
            list(SaleRecord.objects.order_by('external_id').values_list('profit', flat=True)),
            [Decimal('0.50'), Decimal('0.50'), Decimal('30.00')],
        )

    def test_sync_postings_updates_existing_sales_when_finance_appears(self):
        Product.objects.create(
            article='OZON-1',
            name='РўРѕРІР°СЂ Ozon',
            quantity=5,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
        )

        service_without_finance = OzonSyncService(client=FakeOzonClient(finance_transactions=[]))
        first_result = service_without_finance.sync_postings(date(2026, 4, 1), date(2026, 4, 28))
        self.assertEqual(first_result.sales_created, 3)
        self.assertEqual(
            list(SaleRecord.objects.order_by('external_id').values_list('income', flat=True)),
            [Decimal('150.50'), Decimal('150.50'), Decimal('200.00')],
        )

        second_result = self.service.sync_postings(date(2026, 4, 1), date(2026, 4, 28))

        self.assertEqual(second_result.sales_created, 0)
        self.assertEqual(second_result.sales_skipped, 3)
        self.assertEqual(second_result.sales_updated, 3)
        self.assertEqual(
            list(SaleRecord.objects.order_by('external_id').values_list('income', flat=True)),
            [Decimal('120.50'), Decimal('120.50'), Decimal('150.00')],
        )

    def test_sync_postings_splits_long_periods_for_finance_api(self):
        Product.objects.create(
            article='OZON-1',
            name='РўРѕРІР°СЂ Ozon',
            quantity=5,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
        )

        client = FakeOzonClient(finance_transactions=[])
        service = OzonSyncService(client=client)
        service.sync_postings(date(2026, 4, 1), date(2026, 5, 9))

        self.assertGreater(len(client.finance_transaction_periods), 1)
        self.assertEqual(client.finance_transaction_periods[0][0], '2026-04-01T00:00:00Z')
        self.assertEqual(client.finance_transaction_periods[0][1], '2026-04-28T23:59:59.999999Z')
        self.assertEqual(client.finance_transaction_periods[1][0], '2026-04-29T00:00:00Z')


class ProductListTests(TestCase):
    def test_product_list_is_paginated_by_groups(self):
        for index in range(30):
            Product.objects.create(
                article=f'SKU-{index:02d}',
                name=f'Товар {index:02d}',
                quantity=1,
                purchase_price=Decimal('100.00'),
                delivery_cost=Decimal('20.00'),
            )

        first_page = self.client.get(reverse('product_list'))
        second_page = self.client.get(reverse('product_list'), {'page': 2})

        self.assertEqual(first_page.status_code, 200)
        self.assertEqual(second_page.status_code, 200)
        self.assertEqual(len(first_page.context['groups']), 25)
        self.assertEqual(len(second_page.context['groups']), 5)
        self.assertEqual(first_page.context['paginator'].num_pages, 2)
        self.assertEqual(second_page.context['page_obj'].number, 2)

    def test_update_cost_price_can_apply_to_sales_and_undo(self):
        product = Product.objects.create(
            article='SKU-1',
            name='РўРѕРІР°СЂ',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
        )
        sale = SaleRecord.objects.create(
            product=product,
            sale_type='free',
            income=Decimal('200.00'),
        )

        response = self.client.post(reverse('update_cost_price'), {
            'product_id': product.id,
            'cost_price': '150.00',
            'apply_to_sales': 'true',
        })

        self.assertEqual(response.status_code, 200)
        product.refresh_from_db()
        sale.refresh_from_db()
        self.assertEqual(product.cost_price, Decimal('150.00'))
        self.assertEqual(product.purchase_price, Decimal('150.00'))
        self.assertEqual(product.delivery_cost, Decimal('0.00'))
        self.assertEqual(sale.profit, Decimal('50.00'))

        undo_response = self.client.post(reverse('undo_cost_price'))

        self.assertEqual(undo_response.status_code, 200)
        product.refresh_from_db()
        sale.refresh_from_db()
        self.assertEqual(product.cost_price, Decimal('120.00'))
        self.assertEqual(product.purchase_price, Decimal('100.00'))
        self.assertEqual(product.delivery_cost, Decimal('20.00'))
        self.assertEqual(sale.profit, Decimal('80.00'))

    def test_product_list_filters_groups_and_sets_last_sale_date(self):
        positive_product = Product.objects.create(
            article='SKU-POS',
            name='Positive',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='sold',
        )
        negative_product = Product.objects.create(
            article='SKU-NEG',
            name='Negative',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='sold',
        )
        Product.objects.create(
            article='SKU-STOCK',
            name='Stock',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_stock',
        )
        SaleRecord.objects.create(product=positive_product, sale_type='free', income=Decimal('200.00'), sale_date=date(2026, 4, 10))
        SaleRecord.objects.create(product=positive_product, sale_type='free', income=Decimal('210.00'), sale_date=date(2026, 4, 20))
        SaleRecord.objects.create(product=negative_product, sale_type='free', income=Decimal('50.00'), sale_date=date(2026, 4, 15))

        response = self.client.get(reverse('product_list'), {
            'filters_applied': '1',
            'filters': ['sold', 'profit_positive'],
        })

        groups = response.context['groups']
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]['article'], 'SKU-POS')
        self.assertEqual(groups[0]['last_sale_date'], date(2026, 4, 20))

    def test_product_list_can_filter_to_in_sale_only(self):
        Product.objects.create(
            article='SKU-STOCK',
            name='Stock',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_stock',
        )
        Product.objects.create(
            article='SKU-SALE',
            name='Sale',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_sale',
        )

        response = self.client.get(reverse('product_list'), {
            'filters_applied': '1',
            'filters': ['in_sale'],
        })

        groups = response.context['groups']
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]['article'], 'SKU-SALE')

    def test_product_list_shows_in_stock_quantity_as_pieces_without_fake_row_count(self):
        Product.objects.create(
            article='SKU-ZERO',
            name='Zero stock',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_stock',
        )

        response = self.client.get(reverse('product_list'), {'article': 'SKU-ZERO'})

        groups = response.context['groups']
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]['count'], 0)
        self.assertContains(response, '0 штук')
        self.assertNotContains(response, '0 зап.')

    def test_product_list_shows_accrual_fields_for_sold_rows(self):
        product = Product.objects.create(
            article='SKU-SOLD',
            name='Sold',
            quantity=0,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='sold',
        )
        SaleRecord.objects.create(
            product=product,
            sale_type='ozon',
            income=Decimal('200.00'),
            sale_date=date(2026, 4, 10),
            accrual_date=date(2026, 4, 12),
            accrual_id='987654',
            accrual_details={
                'gross_price': '250.00',
                'net_income': '200.00',
                'deductions_total': '50.00',
                'services': [{'name': 'Эквайринг', 'price': '-10.00'}],
                'items': [{'sale_commission': '-40.00'}],
            },
        )

        response = self.client.get(reverse('product_list'), {'article': 'SKU-SOLD'})

        self.assertContains(response, 'Дата начисления')
        self.assertContains(response, 'ID начисления')
        self.assertContains(response, '12.04.2026')
        self.assertContains(response, '987654')
        self.assertContains(response, 'Показать списания')
        self.assertContains(response, 'Эквайринг')

    def test_product_list_can_search_by_article(self):
        Product.objects.create(
            article='JGET100001PERKLIC',
            name='Switch',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_sale',
        )
        Product.objects.create(
            article='SKU-STOCK',
            name='Stock',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_stock',
        )

        response = self.client.get(reverse('product_list'), {'article': 'perklic'})

        groups = response.context['groups']
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]['article'], 'JGET100001PERKLIC')
        self.assertEqual(response.context['active_article_query'], 'perklic')

    def test_product_list_sorts_all_groups_before_pagination(self):
        for index in range(30):
            Product.objects.create(
                article=f'SKU-{index:02d}',
                name=f'Product {index:02d}',
                quantity=1,
                purchase_price=Decimal('100.00'),
                delivery_cost=Decimal('20.00'),
            )

        first_page = self.client.get(reverse('product_list'), {'sort': 'article', 'direction': 'desc'})
        second_page = self.client.get(reverse('product_list'), {'sort': 'article', 'direction': 'desc', 'page': 2})

        self.assertEqual(first_page.context['groups'][0]['article'], 'SKU-29')
        self.assertEqual(first_page.context['groups'][-1]['article'], 'SKU-05')
        self.assertEqual(second_page.context['groups'][0]['article'], 'SKU-04')
        self.assertEqual(second_page.context['groups'][-1]['article'], 'SKU-00')


class SaleUploadTests(TestCase):
    def setUp(self):
        self.product = Product.objects.create(
            article='SKU-1',
            name='Тестовый товар',
            quantity=3,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('25.00'),
        )

    def test_sales_upload_creates_records_and_decreases_stock(self):
        upload = SimpleUploadedFile(
            'sales.csv',
            'Артикул;Название;Доход;Количество\nSKU-1;Тестовый товар;200,50;2\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('upload_sales'),
            {'sale_type': 'ozon', 'file': upload},
        )

        self.assertRedirects(response, reverse('product_list'))
        self.product.refresh_from_db()
        self.assertEqual(self.product.quantity, 1)
        self.assertEqual(self.product.status, 'in_sale')
        self.assertEqual(SaleRecord.objects.count(), 2)
        self.assertEqual(SaleRecord.objects.first().profit, Decimal('75.50'))

    def test_negative_sales_quantity_is_rejected_without_stock_change(self):
        upload = SimpleUploadedFile(
            'sales.csv',
            'Артикул;Название;Доход;Количество\nSKU-1;Тестовый товар;200;-1\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('upload_sales'),
            {'sale_type': 'ozon', 'file': upload},
        )

        self.assertEqual(response.status_code, 200)
        self.product.refresh_from_db()
        self.assertEqual(self.product.quantity, 3)
        self.assertEqual(SaleRecord.objects.count(), 0)

    def test_free_sale_does_not_move_remaining_product_to_in_sale(self):
        upload = SimpleUploadedFile(
            'sales.csv',
            'Артикул;Название;Доход;Количество\nSKU-1;Тестовый товар;200;1\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('upload_sales'),
            {'sale_type': 'free', 'file': upload},
        )

        self.assertRedirects(response, reverse('product_list'))
        self.product.refresh_from_db()
        self.assertEqual(self.product.quantity, 2)
        self.assertEqual(self.product.status, 'in_stock')

    def test_sales_upload_warns_about_name_mismatch_and_oversell(self):
        upload = SimpleUploadedFile(
            'sales.csv',
            'Артикул;Название;Доход;Количество\nSKU-1;Другое имя;200;5\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('upload_sales'),
            {'sale_type': 'free', 'file': upload},
            follow=True,
        )

        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any('название в файле отличается от базы' in message for message in messages))
        self.assertTrue(any('остаток обнулен' in message for message in messages))
        self.product.refresh_from_db()
        self.assertEqual(self.product.status, 'sold')


class SupplyUploadTests(TestCase):
    def test_supply_upload_creates_product_from_csv(self):
        upload = SimpleUploadedFile(
            'supply.CSV',
            'Артикул;Название;Закупка;Доставка;Себестоимость;Количество\nSKU-2;Новый товар;50;10;60;4\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(reverse('upload_supply'), {'file': upload})

        self.assertRedirects(response, reverse('product_list'))
        product = Product.objects.get(article='SKU-2')
        self.assertEqual(product.quantity, 4)
        self.assertEqual(product.cost_price, Decimal('60.00'))

    def test_supply_upload_supports_headerless_file(self):
        upload = SimpleUploadedFile(
            'supply.csv',
            'JGET10022AA;Батарейный блок 2 АА BH325A. J-get;11,38461538;3,415384615;14,8;1880\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(reverse('upload_supply'), {'file': upload})

        self.assertRedirects(response, reverse('product_list'))
        product = Product.objects.get(article='JGET10022AA')
        self.assertEqual(product.quantity, 1880)
        self.assertEqual(product.purchase_price, Decimal('11.38'))
        self.assertEqual(product.delivery_cost, Decimal('3.42'))
        self.assertEqual(product.cost_price, Decimal('14.80'))


class ExportTests(TestCase):
    def test_statistics_page_defaults_to_all_time_stats(self):
        product = Product.objects.create(
            article='SKU-1',
            name='Product',
            quantity=2,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_sale',
        )
        SaleRecord.objects.create(product=product, sale_type='free', income=Decimal('200.00'), sale_date=date(2026, 4, 10))
        SaleRecord.objects.create(product=product, sale_type='free', income=Decimal('50.00'), sale_date=date(2026, 4, 20))

        response = self.client.get(reverse('sales_statistics'))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['has_period'])
        self.assertEqual(response.context['sales_stats_title'], 'За все время')
        self.assertEqual(response.context['sales_stats']['sales_count'], 2)
        self.assertEqual(response.context['sales_stats']['profit'], Decimal('10.00'))
        self.assertEqual(response.context['stock_stats']['in_sale_count'], 2)

    def test_statistics_page_shows_only_period_stats_when_dates_are_selected(self):
        product = Product.objects.create(
            article='SKU-1',
            name='Product',
            quantity=2,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('20.00'),
            status='in_sale',
        )
        SaleRecord.objects.create(product=product, sale_type='free', income=Decimal('200.00'), sale_date=date(2026, 4, 10))
        SaleRecord.objects.create(product=product, sale_type='free', income=Decimal('50.00'), sale_date=date(2026, 4, 20))

        response = self.client.get(reverse('sales_statistics'), {
            'date_from': '2026-04-15',
            'date_to': '2026-04-30',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['has_period'])
        self.assertEqual(response.context['sales_stats_title'], 'За период')
        self.assertEqual(response.context['sales_stats']['sales_count'], 1)
        self.assertEqual(response.context['sales_stats']['profit'], Decimal('-70.00'))
        self.assertEqual(response.context['stock_stats']['in_sale_count'], 2)

    def test_sales_report_period_defaults_date_to_to_today(self):
        response = self.client.get(reverse('sales_report_period'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['form'].initial['date_to'], timezone.localdate())

    def test_sales_report_uses_sale_time_cost_price(self):
        product = Product.objects.create(
            article='SKU-1',
            name='Тестовый товар',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('25.00'),
        )
        SaleRecord.objects.create(product=product, sale_type='free', income=Decimal('200.00'))
        product.purchase_price = Decimal('300.00')
        product.delivery_cost = Decimal('50.00')
        product.save()

        response = self.client.get(reverse('export_sales_report'))
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet['D2'].value, 125)
        self.assertEqual(sheet['F2'].value, 75)

    def test_sales_report_filters_by_period(self):
        product = Product.objects.create(
            article='SKU-1',
            name='Тестовый товар',
            quantity=1,
            purchase_price=Decimal('100.00'),
            delivery_cost=Decimal('25.00'),
        )
        SaleRecord.objects.create(
            product=product,
            sale_type='free',
            income=Decimal('200.00'),
            sale_date=date(2026, 4, 10),
        )
        SaleRecord.objects.create(
            product=product,
            sale_type='free',
            income=Decimal('300.00'),
            sale_date=date(2026, 4, 20),
        )

        response = self.client.get(reverse('export_sales_report'), {
            'date_from': '2026-04-15',
            'date_to': '2026-04-30',
        })
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet['E2'].value, 300)
        self.assertEqual(sheet['G2'].value.date(), date(2026, 4, 20))

    def test_sales_report_rejects_invalid_period(self):
        response = self.client.get(reverse('export_sales_report'), {
            'date_from': '2026-04-30',
            'date_to': '2026-04-01',
        })

        self.assertRedirects(response, reverse('sales_report_period'))

    def test_sales_report_does_not_download_empty_period(self):
        response = self.client.get(reverse('export_sales_report'), {
            'date_from': '2026-04-01',
            'date_to': '2026-04-30',
        }, follow=True)

        self.assertRedirects(response, reverse('sales_report_period'))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertTrue(any('За выбранный период продаж нет' in message for message in messages))


class ParseFileTests(TestCase):
    def test_parse_file_supports_cp1251_csv(self):
        upload = SimpleUploadedFile(
            'sales.csv',
            'Артикул;Название;Доход\nSKU-1;Товар;100\n'.encode('cp1251'),
            content_type='text/csv',
        )

        df = parse_file(upload)

        self.assertEqual(list(df.columns), ['Артикул', 'Название', 'Доход'])
        self.assertEqual(df.iloc[0]['Название'], 'Товар')
