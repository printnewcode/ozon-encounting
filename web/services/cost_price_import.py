import csv
import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.db.models import Count, F, Q
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from web.models import Product, SaleRecord


ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}
HEADER_COST_PREFIXES = ('себ', 'cost')
MAX_COST_PRICE = Decimal('99999999.99')
MAX_FILE_ROWS = 5000


class CostPriceImportError(Exception):
    pass


def normalize_product_name(value) -> str:
    text = unicodedata.normalize('NFKC', str(value or ''))
    return re.sub(r'\s+', ' ', text).strip().casefold()


def is_cost_header(value) -> bool:
    normalized = normalize_product_name(value)
    compact = re.sub(r'[^a-zа-яё]+', '', normalized)
    return any(compact.startswith(prefix) for prefix in HEADER_COST_PREFIXES)


def parse_cost_price(value, row_number: int) -> Decimal:
    if value is None:
        raise CostPriceImportError(f'Строка {row_number}: себестоимость не указана.')

    normalized = str(value).strip().replace('\xa0', '').replace(' ', '').replace('₽', '').replace(',', '.')
    try:
        cost_price = Decimal(normalized).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        raise CostPriceImportError(
            f'Строка {row_number}: некорректная себестоимость «{value}».'
        ) from None

    if cost_price <= 0:
        raise CostPriceImportError(f'Строка {row_number}: себестоимость должна быть больше нуля.')
    if cost_price > MAX_COST_PRICE:
        raise CostPriceImportError(f'Строка {row_number}: себестоимость слишком велика.')
    return cost_price


def read_uploaded_rows(uploaded_file) -> list[tuple]:
    extension = Path(uploaded_file.name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise CostPriceImportError('Поддерживаются только файлы .xlsx и .csv.')

    try:
        if extension == '.xlsx':
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = []
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_number > MAX_FILE_ROWS:
                    raise CostPriceImportError(f'В файле должно быть не больше {MAX_FILE_ROWS} строк.')
                rows.append(tuple(row[:2]))
            return rows

        content = uploaded_file.read()
        decoded = None
        for encoding in ('utf-8-sig', 'cp1251'):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise CostPriceImportError('Не удалось определить кодировку CSV. Используйте UTF-8 или CP1251.')

        sample = decoded[:4096]
        delimiter = ';' if ';' in sample else ','
        rows = []
        for row_number, row in enumerate(csv.reader(io.StringIO(decoded), delimiter=delimiter), start=1):
            if row_number > MAX_FILE_ROWS:
                raise CostPriceImportError(f'В файле должно быть не больше {MAX_FILE_ROWS} строк.')
            rows.append(tuple(row[:2]))
        return rows
    except (InvalidFileException, OSError, ValueError) as exc:
        raise CostPriceImportError(f'Не удалось прочитать файл: {exc}') from exc


def parsed_cost_rows(uploaded_file) -> list[dict]:
    raw_rows = read_uploaded_rows(uploaded_file)
    rows = []

    for row_number, values in enumerate(raw_rows, start=1):
        name = values[0] if len(values) > 0 else None
        cost_value = values[1] if len(values) > 1 else None
        if not str(name or '').strip() and not str(cost_value or '').strip():
            continue

        normalized_name = normalize_product_name(name)
        if not rows and is_cost_header(cost_value):
            continue

        if not normalized_name:
            rows.append({
                'row_number': row_number,
                'name': '',
                'cost_price': str(cost_value or ''),
                'status': 'invalid',
                'message': 'Название товара не указано.',
            })
            continue

        try:
            cost_price = parse_cost_price(cost_value, row_number)
        except CostPriceImportError as exc:
            rows.append({
                'row_number': row_number,
                'name': str(name).strip(),
                'cost_price': str(cost_value or ''),
                'status': 'invalid',
                'message': str(exc),
            })
            continue

        rows.append({
            'row_number': row_number,
            'name': str(name).strip(),
            'normalized_name': normalized_name,
            'cost_price': str(cost_price),
        })

    if not rows:
        raise CostPriceImportError('В файле нет строк с товарами.')
    return rows


def build_import_preview(uploaded_file) -> dict:
    rows = parsed_cost_rows(uploaded_file)
    costs_by_name = {}
    for row in rows:
        normalized_name = row.get('normalized_name')
        if normalized_name:
            costs_by_name.setdefault(normalized_name, set()).add(row['cost_price'])
    conflicting_duplicate_names = {
        name for name, cost_prices in costs_by_name.items() if len(cost_prices) > 1
    }

    candidates = Product.objects.annotate(
        batch_count=Count('batches', distinct=True),
        zero_ozon_sales_count=Count(
            'sales',
            filter=Q(sales__sale_type='ozon') & (
                Q(sales__cost_price=Decimal('0.00')) | Q(sales__cost_price__isnull=True)
            ),
            distinct=True,
        )
    )
    candidates_by_name = {}
    for product in candidates:
        product.update_product_cost = (
            product.ozon_quantity > 0
            and product.quantity == 0
            and product.cost_price == Decimal('0.00')
            and product.batch_count == 0
        )
        if product.update_product_cost or product.zero_ozon_sales_count > 0:
            candidates_by_name.setdefault(normalize_product_name(product.name), []).append(product)

    all_products_by_name = {}
    for product in Product.objects.only('id', 'article', 'name'):
        all_products_by_name.setdefault(normalize_product_name(product.name), []).append(product)

    preview_rows = []
    seen_product_names = set()
    for row in rows:
        if row.get('status') == 'invalid':
            preview_rows.append(row)
            continue

        normalized_name = row['normalized_name']
        if normalized_name in conflicting_duplicate_names:
            row.update(
                status='duplicate',
                message='Для одного названия указана разная себестоимость.',
            )
            preview_rows.append(row)
            continue
        if normalized_name in seen_product_names:
            row.update(
                status='skipped_duplicate',
                message='Дубликат с той же себестоимостью пропущен.',
            )
            preview_rows.append(row)
            continue
        seen_product_names.add(normalized_name)

        matches = candidates_by_name.get(normalized_name, [])
        if len(matches) == 1:
            product = matches[0]
            actions = []
            if product.update_product_cost:
                actions.append('карточка товара')
            if product.zero_ozon_sales_count:
                actions.append(f'прошлых продаж: {product.zero_ozon_sales_count}')
            row.update(
                status='matched',
                message=f'Будет обновлено: {", ".join(actions)}.',
                product_id=product.id,
                article=product.article,
                update_product=product.update_product_cost,
                past_sales_count=product.zero_ozon_sales_count,
                old_purchase_price=str(product.purchase_price),
                old_delivery_cost=str(product.delivery_cost),
                old_cost_price=str(product.cost_price),
            )
        elif len(matches) > 1:
            row.update(
                status='ambiguous',
                message='Найдено несколько подходящих OZON-товаров.',
            )
        elif normalized_name in all_products_by_name:
            articles = ', '.join(product.article for product in all_products_by_name[normalized_name][:5])
            row.update(
                status='protected',
                article=articles,
                message='Товар найден, но защищён условиями импорта.',
            )
        else:
            row.update(
                status='not_found',
                message='Товар с таким названием не найден.',
            )
        preview_rows.append(row)

    summary = {
        status: sum(row.get('status') == status for row in preview_rows)
        for status in (
            'matched',
            'protected',
            'not_found',
            'ambiguous',
            'duplicate',
            'skipped_duplicate',
            'invalid',
        )
    }
    summary['products_to_update'] = sum(
        bool(row.get('update_product')) for row in preview_rows if row.get('status') == 'matched'
    )
    summary['past_sales_to_update'] = sum(
        int(row.get('past_sales_count') or 0) for row in preview_rows if row.get('status') == 'matched'
    )
    return {
        'file_name': Path(uploaded_file.name).name[:255],
        'rows': preview_rows,
        'summary': summary,
    }


def product_is_import_target(product: Product) -> bool:
    return (
        product.ozon_quantity > 0
        and product.quantity == 0
        and product.cost_price == Decimal('0.00')
        and not product.batches.exists()
    )


@transaction.atomic
def apply_cost_price_import(preview: dict, apply_to_past_sales: bool = False) -> list[dict]:
    matched_rows = [row for row in preview.get('rows', []) if row.get('status') == 'matched']
    if not matched_rows:
        raise CostPriceImportError('Нет товаров, готовых к обновлению.')
    if not apply_to_past_sales and not any(row.get('update_product') for row in matched_rows):
        raise CostPriceImportError('Включите обновление прошлых продаж или загрузите другой файл.')

    products = {
        product.id: product
        for product in Product.objects.select_for_update().filter(
            id__in=[row['product_id'] for row in matched_rows]
        )
    }
    for row in matched_rows:
        product = products.get(row['product_id'])
        if (
            product is None
            or normalize_product_name(product.name) != row['normalized_name']
        ):
            raise CostPriceImportError(
                f'Товар «{row["name"]}» изменился после предпросмотра. Создайте импорт заново.'
            )
        if row.get('update_product') and (
            str(product.purchase_price) != row['old_purchase_price']
            or str(product.delivery_cost) != row['old_delivery_cost']
            or str(product.cost_price) != row['old_cost_price']
            or not product_is_import_target(product)
        ):
            raise CostPriceImportError(
                f'Товар «{row["name"]}» изменился после предпросмотра. Создайте импорт заново.'
            )

    results = []
    for row in matched_rows:
        product = products[row['product_id']]
        new_cost_price = Decimal(row['cost_price'])
        product_updated = False
        if row.get('update_product'):
            product.purchase_price = new_cost_price
            product.delivery_cost = Decimal('0.00')
            product.save()
            product_updated = True

        sales_updated = 0
        if apply_to_past_sales:
            sale_ids = list(
                SaleRecord.objects.select_for_update()
                .filter(product=product, sale_type='ozon')
                .filter(Q(cost_price=Decimal('0.00')) | Q(cost_price__isnull=True))
                .values_list('id', flat=True)
            )
            if sale_ids:
                sales_updated = SaleRecord.objects.filter(id__in=sale_ids).update(
                    cost_price=new_cost_price,
                    profit=F('income') - new_cost_price,
                )

        if product_updated or sales_updated:
            results.append({
                'product': product,
                'row': row,
                'product_updated': product_updated,
                'sales_updated': sales_updated,
            })

    if not results:
        raise CostPriceImportError('Нет данных, которые можно обновить.')
    return results
